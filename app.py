import csv
import io
import os
import sys
import oracledb
import ldap3
import oracledb
import json
from collections import Counter
from datetime import date, datetime, timedelta
from math import ceil
from time import time
from urllib.parse import quote_plus
from dotenv import load_dotenv
from flask import Flask, Response, current_app, g, jsonify, redirect, render_template, request, session
from flask_socketio import SocketIO, emit, join_room, leave_room
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Column, DateTime, Integer, String, create_engine, func, or_
from sqlalchemy.orm import declarative_base, sessionmaker
from werkzeug.security import check_password_hash
from sqlalchemy.exc import IntegrityError

load_dotenv()
usuarios_online = {}

os.environ["LD_LIBRARY_PATH"] = "/opt/oracle/instantclient"
oracledb.init_oracle_client(
    lib_dir="/opt/oracle/instantclient"
)

CACHE_PUNTOS = {}
CACHE_TTL = 60

def get_oracle_points(idcliente):

    if not oracle_pool:
        return 0

    conn = None
    cursor = None

    try:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT NVL(SUM(PUNTOS),0)
            FROM SISTSIO.RIOREWARDS_DETALLE
            WHERE IDCLIENTE = :idcliente
        """, {
            "idcliente": int(idcliente)
        })

        row = cursor.fetchone()

        return int(row[0] or 0)

    except Exception as exc:

        print("ORACLE POINTS ERROR:", exc)
        return 0

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

def insertar_puntos_oracle(
    idcliente,
    idcita,
    cantidad,
    categoria,
    usuario
):
    if not oracle_pool:
        return

    conn = None
    cursor = None

    try:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO SISTSIO.RIOREWARDS_DETALLE
            (
                IDCLIENTE,
                IDCITA,
                CANTIDAD,
                PUNTOS,
                TIPORECOMPENSA,
                COMENTARIO,
                CORREOREGISTRADO
            )
            VALUES
            (
                :idcliente,
                :idcita,
                0,
                :puntos,
                :tipo,
                :comentario,
                NULL
            )
        """, {
            "idcliente": int(idcliente),
            "idcita": int(idcita) if idcita else None,
            "puntos": int(cantidad),
            "tipo": "Reseña",
            "comentario": "Bonificación por reseña"
        })

        conn.commit()

        print(
            f"ORACLE OK -> cliente={idcliente} puntos={cantidad}"
        )

    except Exception as exc:

        print("ERROR INSERT ORACLE:", exc)

        raise

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

def insertar_cancelacion_resena_oracle(
    idcliente,
    idcita,
    cantidad,
    usuario
):
    if not oracle_pool:
        return

    conn = None
    cursor = None

    try:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO SISTSIO.RIOREWARDS_DETALLE
            (
                IDCLIENTE,
                IDCITA,
                CANTIDAD,
                PUNTOS,
                TIPORECOMPENSA,
                COMENTARIO,
                CORREOREGISTRADO
            )
            VALUES
            (
                :idcliente,
                :idcita,
                0,
                :puntos,
                :tipo,
                :comentario,
                NULL
            )
        """, {
            "idcliente": int(idcliente),
            "idcita": int(idcita) if idcita else None,
            "puntos": -abs(int(cantidad)),
            "tipo": "Cancelación reseña",
            "comentario": "Cancelación de bonificación por reseña"
        })

        conn.commit()

    except Exception as exc:

        print("ERROR CANCELANDO RESEÑA:", exc)
        raise

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

def paciente_tiene_resena_oracle(idcliente):

    if not oracle_pool:
        return False

    conn = None
    cursor = None

    try:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT NVL(SUM(PUNTOS),0)
            FROM SISTSIO.RIOREWARDS_DETALLE
            WHERE IDCLIENTE = :idcliente
            AND UPPER(TIPORECOMPENSA) IN (
                UPPER('Reseña'),
                UPPER('Cancelación reseña')
            )
        """, {
            "idcliente": int(idcliente)
        })

        row = cursor.fetchone()

        saldo_resena = int(row[0] or 0)

        return saldo_resena > 0

    except Exception as exc:

        print("ERROR VALIDANDO RESEÑA:", exc)
        return False

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

def tiene_resena_oracle(idcliente):

    if not oracle_pool:
        return False

    conn = None
    cursor = None

    try:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT NVL(SUM(PUNTOS),0)
            FROM SISTSIO.RIOREWARDS_DETALLE
            WHERE IDCLIENTE = :idcliente
              AND UPPER(TIPORECOMPENSA) IN (
                    UPPER('Reseña'),
                    UPPER('Cancelación reseña')
              )
        """, {
            "idcliente": int(idcliente)
        })

        saldo = int(cursor.fetchone()[0] or 0)

        return saldo > 0

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

def build_engine():
    db_user = os.getenv("DB_USER")
    db_password = quote_plus(os.getenv("DB_PASSWORD") or "")
    db_host = os.getenv("DB_HOST")
    db_name = os.getenv("DB_NAME")

    if db_user and db_host and db_name:
        return create_engine(
            f"postgresql+psycopg2://{db_user}:{db_password}@{db_host}:5432/{db_name}",
            pool_pre_ping=True,
            pool_recycle=1800,
            pool_size=10,
            max_overflow=20,
            pool_timeout=30,
            echo=False,
        )

    base_dir = os.path.dirname(os.path.abspath(__file__))
    sqlite_path = os.path.join(base_dir, "database", "sistema.db")
    return create_engine(f"sqlite:///{sqlite_path}", future=True)

ORACLE_USER = os.getenv("ORACLE_USER")
ORACLE_PASSWORD = os.getenv("ORACLE_PASSWORD")
ORACLE_HOST = os.getenv("ORACLE_HOST")
ORACLE_PORT = os.getenv("ORACLE_PORT")
ORACLE_SERVICE = os.getenv("ORACLE_SERVICE")

def build_oracle_pool():

    try:

        dsn = f"{ORACLE_HOST}:{ORACLE_PORT}/{ORACLE_SERVICE}"

        pool = oracledb.create_pool(
            user=ORACLE_USER,
            password=ORACLE_PASSWORD,
            dsn=dsn,
            min=1,
            max=5,
            increment=1
        )

        return pool

    except Exception as e:

        print("ERROR CREANDO POOL ORACLE:")
        print(str(e))

        return None


engine = build_engine()
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()
oracle_pool = build_oracle_pool()
print("ORACLE POOL CREADO:", oracle_pool)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY") or "clave_super_secreta"
socketio = SocketIO(app, cors_allowed_origins="*",async_mode="gevent") # agregar dominio ahorita slo tiene un simbolo

LDAP_SERVER = os.getenv("LDAP_SERVER")
LDAP_BASE_DN = os.getenv("LDAP_BASE_DN")
LDAP_ADMIN_USER = os.getenv("LDAP_USER_DN")
LDAP_ADMIN_PASSWORD = os.getenv("LDAP_PASSWORD")
SUPERADMIN_USER = os.getenv("SUPERADMIN_USER")
SUPERADMIN_PASSWORD_HASH = os.getenv("SUPERADMIN_PASSWORD_HASH")


class Usuario(Base):
    __tablename__ = "usuarios"

    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True)
    rol = Column(String(50))
    permisos = Column(String(200))


class Permiso(Base):
    __tablename__ = "permisos"

    id = Column(Integer, primary_key=True)
    nombre = Column(String(50), unique=True)


class Rol(Base):
    __tablename__ = "roles"

    id = Column(Integer, primary_key=True)
    nombre = Column(String(50), unique=True)
    permisos = Column(String(200))


class Auditoria(Base):
    __tablename__ = "auditoria"

    id = Column(Integer, primary_key=True)
    usuario_admin = Column(String(50))
    accion = Column(String(50))
    usuario_afectado = Column(String(50))
    detalle = Column(String(200))
    fecha = Column(DateTime, default=datetime.now)


class Notificacion(Base):
    __tablename__ = "notificaciones"

    id = Column(Integer, primary_key=True)
    mensaje = Column(String(200))
    usuario = Column(String(50))
    leido = Column(Integer, default=0)
    fecha = Column(DateTime, default=datetime.now)


class Puntos(Base):
    __tablename__ = "puntos"

    id = Column(Integer, primary_key=True)
    paciente_id = Column(String(50), unique=True)
    puntos = Column(Integer, default=0)
    actualizado_en = Column(DateTime, default=datetime.now)


class MovimientoPuntos(Base):
    __tablename__ = "movimientos_puntos"

    id = Column(Integer, primary_key=True)
    paciente_id = Column(String(50))
    idcita = Column(String(50))
    tipo = Column(String(20))
    cantidad = Column(Integer)
    usuario_admin = Column(String(50))
    fecha = Column(DateTime, default=datetime.now)
    motivo = Column(String(50))
    origen = Column(String(50))


def get_db():
    if not hasattr(g, "db") or g.db is None:
        g.db = SessionLocal()
    return g.db


def get_usuario():
    return session.get("usuario")


def serialize_datetime(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value is None:
        return ""
    return str(value)


def parse_date_param(value, end_of_day=False):
    if not value:
        return None
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
        if end_of_day:
            return parsed + timedelta(days=1) - timedelta(seconds=1)
        return parsed
    except ValueError:
        return None


def get_report_filters(source):
    return {
        "tipo": (source.get("tipo") or "puntos").strip().lower(),

        "usuario": (source.get("usuario") or "").strip(),

        "fecha_inicio": (
            source.get("fecha_inicio") or ""
        ).strip(),

        "fecha_fin": (
            source.get("fecha_fin") or ""
        ).strip(),

        "movimiento_tipo": (
            source.get("movimiento_tipo") or ""
        ).strip().lower(),

        # 🔥 NUEVO
        "pagina": int(source.get("pagina", 1)),

        # 🔥 NUEVO
        "por_pagina": int(source.get("por_pagina", 20)),
    }


def serialize_movimiento(registro):
    return {
        "id": registro.id,
        "paciente_id": str(registro.paciente_id),
        "idcita": registro.idcita or "",
        "tipo": registro.tipo or "",
        "cantidad": int(registro.cantidad or 0),
        "usuario_admin": registro.usuario_admin or "",
        "fecha": serialize_datetime(registro.fecha),
        "motivo": registro.motivo or "",
    }


def serialize_auditoria(registro):
    return {
        "id": registro.id,
        "usuario_admin": registro.usuario_admin or "",
        "accion": registro.accion or "",
        "usuario_afectado": registro.usuario_afectado or "",
        "detalle": registro.detalle or "",
        "fecha": serialize_datetime(registro.fecha),
    }


def build_filtered_movimientos_query(db, filtros):
    query = db.query(MovimientoPuntos)
    usuario = filtros.get("usuario")
    fecha_inicio = parse_date_param(filtros.get("fecha_inicio"))
    fecha_fin = parse_date_param(filtros.get("fecha_fin"), end_of_day=True)
    movimiento_tipo = filtros.get("movimiento_tipo")

    if usuario:
        query = query.filter(MovimientoPuntos.usuario_admin.ilike(f"%{usuario}%"))
    if fecha_inicio:
        query = query.filter(MovimientoPuntos.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(MovimientoPuntos.fecha <= fecha_fin)
    if movimiento_tipo in {"suma", "resta"}:
        query = query.filter(MovimientoPuntos.tipo == movimiento_tipo)

    return query


def build_filtered_auditoria_query(db, filtros):
    query = db.query(Auditoria)
    usuario = filtros.get("usuario")
    fecha_inicio = parse_date_param(filtros.get("fecha_inicio"))
    fecha_fin = parse_date_param(filtros.get("fecha_fin"), end_of_day=True)

    if usuario:
        query = query.filter(Auditoria.usuario_admin.ilike(f"%{usuario}%"))
    if fecha_inicio:
        query = query.filter(Auditoria.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Auditoria.fecha <= fecha_fin)

    return query


def obtener_historial_oracle(idcliente):

    historial = []

    # ======================
    # HISTORIAL ORACLE
    # ======================
    if oracle_pool:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                PUNTOS,
                TIPORECOMPENSA,
                COMENTARIO,
                FECHATRANSACCION
            FROM SISTSIO.RIOREWARDS_DETALLE
            WHERE IDCLIENTE = :idcliente
        """, {
            "idcliente": int(idcliente)
        })

        rows = cursor.fetchall()

        for row in rows:

            puntos = int(row[0] or 0)

            historial.append({
                "tipo": "suma" if puntos >= 0 else "resta",
                "cantidad": abs(puntos),
                "motivo": row[1] or "",
                "comentario": row[2] or "",
                "usuario_admin": "Oracle",
                "fecha": (
                    row[3].strftime("%Y-%m-%d %H:%M:%S")
                    if row[3]
                    else ""
                )
            })

        cursor.close()
        conn.close()

    # ======================
    # HISTORIAL LOCAL
    # ======================
    db = get_db()

    movimientos = (
        db.query(MovimientoPuntos)
        .filter(
            MovimientoPuntos.paciente_id == str(idcliente)
        )
        .all()
    )

    for mov in movimientos:

        if (mov.motivo or "").lower() == "reseña":
            continue

        historial.append({
            "tipo": mov.tipo,
            "cantidad": mov.cantidad,
            "motivo": mov.motivo or "",
            "comentario": mov.motivo or "",
            "usuario_admin": mov.usuario_admin or "",
            "fecha": (
                mov.fecha.strftime("%Y-%m-%d %H:%M:%S")
                if mov.fecha
                else ""
            )
        })

    historial.sort(
        key=lambda x: x["fecha"],
        reverse=True
    )

    return historial

        

def calcular_saldo_local(db, idcliente):

    movimientos = (
        db.query(MovimientoPuntos)
        .filter(
            MovimientoPuntos.paciente_id == str(idcliente)
        )
        .all()
    )

    total = 0

    for mov in movimientos:

        # IGNORAR cualquier movimiento de reseña
        if (mov.motivo or "").lower() == "reseña":
            continue

        if mov.tipo == "suma":
            total += mov.cantidad

        elif mov.tipo == "resta":
            total -= mov.cantidad

    return total

def get_operational_points(db, idcliente):

    oracle_points = get_oracle_points(idcliente)
    local_delta = calcular_saldo_local(db, idcliente)

    return {
        "oracle": oracle_points,
        "ajustes_locales": local_delta,
        "total": (oracle_points or 0) + (local_delta or 0),
    }


def tiene_resena(idcliente):
    if not oracle_pool:
        return False

    conn = None
    cursor = None
    try:
        conn = oracle_pool.acquire()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT COUNT(*)
            FROM SISTSIO.RIOREWARDS_DETALLE
            WHERE idcliente = :idcliente
              AND UPPER(TIPORECOMPENSA) LIKE '%RESE%'
            """,
            {"idcliente": idcliente},
        )
        count = cursor.fetchone()[0]
        return count > 0
    except Exception as exc:
        print("ORACLE RESENA ERROR:", exc)
        return False
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def ya_tiene_resena_local(db, idcliente):
    existe = db.query(MovimientoPuntos).filter(
        MovimientoPuntos.paciente_id == str(idcliente),
        MovimientoPuntos.motivo == "reseña"
    ).first()
    return existe is not None


def build_report_context(db, filtros):
    pagina = int(filtros.get("pagina", 1))
    por_pagina = int(
        filtros.get("por_pagina", 8)
    )
    offset = (pagina - 1) * por_pagina

    query_movimientos = build_filtered_movimientos_query(db, filtros)

    total_movimientos = query_movimientos.count()

    movimientos_chart = (
        query_movimientos
        .order_by(
            MovimientoPuntos.fecha.desc(),
            MovimientoPuntos.id.desc()
        )
        .all()
    )

    movimientos = (
        query_movimientos
        .order_by(
            MovimientoPuntos.fecha.desc(),
            MovimientoPuntos.id.desc()
        )
        .offset(offset)
        .limit(por_pagina)
        .all()
    )
    query_auditoria = build_filtered_auditoria_query(db, filtros)

    auditoria_chart = (
        query_auditoria
        .order_by(
            Auditoria.fecha.desc(),
            Auditoria.id.desc()
        )
        .all()
    )

    auditoria = (
        query_auditoria
        .order_by(
            Auditoria.fecha.desc(),
            Auditoria.id.desc()
        )
        .offset(offset)
        .limit(por_pagina)
        .all()
    )
    usuarios = db.query(Usuario).order_by(
        Usuario.username.asc()
    ).all()

    puntos_por_fecha = Counter()
    movimientos_por_fecha = Counter()
    movimientos_por_admin = Counter()
    for movimiento in movimientos_chart:

        fecha_key = serialize_datetime(
            movimiento.fecha
        ).split(" ")[0]

        valor = (
            movimiento.cantidad
            if movimiento.tipo == "suma"
            else -movimiento.cantidad
        )

        puntos_por_fecha[fecha_key] += valor

        movimientos_por_fecha[fecha_key] += 1

        movimientos_por_admin[
            movimiento.usuario_admin
            or "Sin usuario"
        ] += 1

    auditoria_por_accion = Counter(
        registro.accion or "Sin accion"
        for registro in auditoria_chart
    )
    auditoria_por_fecha = Counter(
        serialize_datetime(registro.fecha).split(" ")[0]
        for registro in auditoria_chart
        if registro.fecha
    )
    roles = Counter((usuario.rol or "usuario") for usuario in usuarios)

    chart_roles = {
        "labels": ["Admin", "Usuario", "Superadmin"],
        "valores": [
            roles.get("admin", 0),
            roles.get("usuario", 0),
            roles.get("superadmin", 0),
        ],
    }
    chart_puntos_fecha = {
        "labels": list(puntos_por_fecha.keys()),
        "valores": [int(valor) for valor in puntos_por_fecha.values()],
    }
    chart_movimientos_fecha = {
        "labels": list(movimientos_por_fecha.keys()),
        "valores": [int(valor) for valor in movimientos_por_fecha.values()],
    }
    chart_movimientos_admin = {
        "labels": list(movimientos_por_admin.keys()),
        "valores": [int(valor) for valor in movimientos_por_admin.values()],
    }
    chart_auditoria = {
        "labels": list(auditoria_por_accion.keys()),
        "valores": [int(valor) for valor in auditoria_por_accion.values()],
    }
    chart_auditoria_fecha = {
        "labels": list(auditoria_por_fecha.keys()),
        "valores": [int(valor) for valor in auditoria_por_fecha.values()],
    }

    datos_puntos = [serialize_movimiento(movimiento) for movimiento in movimientos]
    datos_admin = [serialize_auditoria(registro) for registro in auditoria]
    tipo = filtros.get("tipo", "puntos")

    kpis = {
        "usuarios": len(usuarios),

        "admins":
            roles.get("admin", 0)
            + roles.get("superadmin", 0),

        "movimientos_puntos":
            total_movimientos,

        "auditoria":
            len(auditoria),

        "puntos_netos":
            len(auditoria)
            if tipo == "admin"
            else sum(
                mov.cantidad
                if mov.tipo == "suma"
                else -mov.cantidad
                for mov in db.query(MovimientoPuntos).all()
            ),

        "movimientos_hoy":
            len(auditoria_chart)
            if tipo == "admin"
            else sum(
                1
                for registro in datos_puntos
                if registro["fecha"].startswith(
                    date.today().isoformat()
                )
            )
    }

    if tipo == "admin":

        chart_principal = {
            "labels": chart_auditoria["labels"],
            "valores": chart_auditoria["valores"]
        }

        chart_fechas = {
            "labels": chart_auditoria_fecha["labels"],
            "valores": chart_auditoria_fecha["valores"]
        }

    else:

        chart_principal = {
            "labels": chart_movimientos_admin["labels"],
            "valores": chart_movimientos_admin["valores"]
        }

        chart_fechas = {
            "labels": chart_puntos_fecha["labels"],
            "valores": chart_puntos_fecha["valores"]
        }
    return {
        "tipo": filtros.get("tipo", "puntos"),
        "filtros": filtros,
        "datos_puntos": datos_puntos,
        "datos_admin": datos_admin,
        "kpis": kpis,
        "pagina": pagina,
        "por_pagina": por_pagina,
        "total_movimientos": total_movimientos,
        "chart_roles": chart_roles,
        "chart_puntos_fecha": chart_puntos_fecha,
        "chart_movimientos_fecha": chart_movimientos_fecha,
        "chart_movimientos_admin": chart_movimientos_admin,
        "chart_auditoria": chart_auditoria,
        "chart_auditoria_fecha": chart_auditoria_fecha,
        "labels": chart_roles["labels"],
        "valores": chart_roles["valores"],
        "valores_admin": [roles.get("admin", 0)],
        "valores_user": [roles.get("usuario", 0)],
        "valores_super": [roles.get("superadmin", 0)],
        "chart_principal": chart_principal,
        "chart_fechas": chart_fechas,
    }


def exportar_reportes_excel(context):
    workbook = Workbook()
    resumen = workbook.active
    resumen.title = "Resumen"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_sheet(sheet):
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = width + 2

    resumen.append(["Metrica", "Valor"])
    resumen.append(["Usuarios", context["kpis"]["usuarios"]])
    resumen.append(["Admins", context["kpis"]["admins"]])
    resumen.append(["Movimientos puntos", context["kpis"]["movimientos_puntos"]])
    resumen.append(["Puntos netos", context["kpis"]["puntos_netos"]])
    resumen.append(["Registros auditoria", context["kpis"]["auditoria"]])
    style_sheet(resumen)

    puntos_sheet = workbook.create_sheet("Puntos")
    puntos_sheet.append(["Paciente", "IDCita", "Tipo", "Cantidad", "Admin", "Fecha"])
    for row in context["datos_puntos"]:
        puntos_sheet.append([
            row["paciente_id"],
            row["idcita"],
            row["tipo"],
            row["cantidad"],
            row["usuario_admin"],
            row["fecha"],
        ])
    style_sheet(puntos_sheet)

    auditoria_sheet = workbook.create_sheet("Auditoria")
    auditoria_sheet.append(["Admin", "Accion", "Usuario", "Detalle", "Fecha"])
    for row in context["datos_admin"]:
        auditoria_sheet.append([
            row["usuario_admin"],
            row["accion"],
            row["usuario_afectado"],
            row["detalle"],
            row["fecha"],
        ])
    style_sheet(auditoria_sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_sistema.xlsx"},
    )


def exportar_reportes_csv(context):
    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["seccion", "campo_1", "campo_2", "campo_3", "campo_4", "campo_5", "campo_6"])
    for row in context["datos_puntos"]:
        writer.writerow([
            "puntos",
            row["paciente_id"],
            row["idcita"],
            row["tipo"],
            row["cantidad"],
            row["usuario_admin"],
            row["fecha"],
        ])
    for row in context["datos_admin"]:
        writer.writerow([
            "auditoria",
            row["usuario_admin"],
            row["accion"],
            row["usuario_afectado"],
            row["detalle"],
            row["fecha"],
            "",
        ])
    return Response(
        stream.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=reporte_sistema.csv"},
    )


def crear_base(db):
    permisos_base = ["dashboard", "usuarios", "reportes", "puntos"]
    roles_base = {
        "usuario": "dashboard",
        "admin": "dashboard,usuarios,reportes,puntos",
        "superadmin": "all",
    }

    for permiso in permisos_base:
        if not db.query(Permiso).filter_by(nombre=permiso).first():
            db.add(Permiso(nombre=permiso))

    for nombre, permisos in roles_base.items():
        rol = db.query(Rol).filter_by(nombre=nombre).first()
        if not rol:
            db.add(Rol(nombre=nombre, permisos=permisos))
        elif nombre != "superadmin" and "puntos" not in (rol.permisos or ""):
            rol.permisos = permisos

    db.commit()


def autenticar_ldap(username, password):
    if not all([LDAP_SERVER, LDAP_BASE_DN, username, password]):
        return False
    try:
        server = ldap3.Server(LDAP_SERVER)
        conn = ldap3.Connection(
            server,
            user=f"uid={username},{LDAP_BASE_DN}",
            password=password,
            auto_bind=True,
        )
        conn.unbind()
        return True
    except Exception as exc:
        print("LDAP LOGIN ERROR:", exc)
        return False


def obtener_usuarios_ldap():
    if not all([LDAP_SERVER, LDAP_BASE_DN, LDAP_ADMIN_USER, LDAP_ADMIN_PASSWORD]):
        return []
    try:
        server = ldap3.Server(LDAP_SERVER)
        conn = ldap3.Connection(
            server,
            user=LDAP_ADMIN_USER,
            password=LDAP_ADMIN_PASSWORD,
            auto_bind=True,
        )
        conn.search(LDAP_BASE_DN, "(objectClass=person)", attributes=["uid"])
        usuarios = sorted({entry.uid.value for entry in conn.entries if hasattr(entry, "uid")})
        conn.unbind()
        return usuarios
    except Exception as exc:
        print("LDAP USERS ERROR:", exc)
        return []


def obtener_grupos_ldap(username):
    if not all([LDAP_SERVER, LDAP_BASE_DN, LDAP_ADMIN_USER, LDAP_ADMIN_PASSWORD, username]):
        return []
    try:
        server = ldap3.Server(LDAP_SERVER)
        conn = ldap3.Connection(
            server,
            user=LDAP_ADMIN_USER,
            password=LDAP_ADMIN_PASSWORD,
            auto_bind=True,
        )
        conn.search(LDAP_BASE_DN, f"(uid={username})", attributes=["memberOf"])
        grupos = []
        for entry in conn.entries:
            if "memberOf" in entry:
                grupos = list(entry.memberOf.values)
        conn.unbind()
        return grupos
    except Exception as exc:
        print("LDAP GROUP ERROR:", exc)
        return []


def mapear_rol_por_grupo(grupos):
    mapa = {
        "cn=admins": "admin",
        "cn=sistemas": "admin",
        "cn=rh": "usuario",
    }
    for grupo in grupos:
        grupo_normalizado = grupo.lower()
        for clave, rol in mapa.items():
            if clave in grupo_normalizado:
                return rol
    return "usuario"


def sync_ldap(db):
    for username in obtener_usuarios_ldap():
        if not db.query(Usuario).filter_by(username=username).first():
            db.add(Usuario(username=username, rol="usuario", permisos=""))
    db.commit()


def log_auditoria(admin, accion, usuario, detalle):
    db = get_db()
    db.add(
        Auditoria(
            usuario_admin=admin,
            accion=accion,
            usuario_afectado=usuario,
            detalle=detalle,
            fecha=datetime.now(),
        )
    )
    db.commit()


def requiere_permiso(permiso):
    def decorator(funcion):
        def wrapped(*args, **kwargs):
            db = get_db()
            if "usuario" not in session:
                return redirect("/")

            usuario = db.query(Usuario).filter_by(username=session["usuario"]).first()
            if not usuario:
                return redirect("/")

            if usuario.rol == "superadmin":
                return funcion(*args, **kwargs)

            rol = db.query(Rol).filter_by(nombre=usuario.rol).first()
            permisos_rol = (rol.permisos or "").split(",") if rol else []
            permisos_usuario = (usuario.permisos or "").split(",")
            permisos_finales = {valor.strip() for valor in permisos_rol + permisos_usuario if valor}

            if permiso not in permisos_finales:
                return "No tienes permiso", 403
            return funcion(*args, **kwargs)

        wrapped.__name__ = funcion.__name__
        return wrapped

    return decorator


def obtener_usuarios_ldap_cache():
    cache = current_app.config.setdefault("LDAP_CACHE", {"data": [], "last_update": None})
    last_update = cache.get("last_update")
    if not last_update or datetime.now() - last_update > timedelta(minutes=2):
        cache["data"] = obtener_usuarios_ldap()
        cache["last_update"] = datetime.now()
    return cache["data"]


def get_unread_notifications_count(db, username):
    if not username:
        return 0
    return db.query(Notificacion).filter_by(usuario=username, leido=0).count()


try:
    Base.metadata.create_all(engine)
    db_bootstrap = SessionLocal()
    crear_base(db_bootstrap)
    db_bootstrap.close()
except Exception as exc:
    print("DATABASE INIT ERROR:", exc)


@app.before_request
def before_request():
    if not hasattr(g, "db"):
        g.db = SessionLocal()


@app.teardown_appcontext
def shutdown_session(exception=None):
    db = getattr(g, "db", None)
    if db is not None:
        try:
            if exception:
                db.rollback()
            db.close()
        except Exception as exc:
            print("DB CLOSE ERROR:", exc)


@app.route("/", methods=["GET", "POST"])
def login():
    db = get_db()
    error = None

    if request.method == "POST":
        username = request.form.get("username") or request.form.get("usuario")
        password = request.form.get("password")

        if not username or not password:
            error = "Faltan datos"
            return render_template("login.html", error=error)

        if username == SUPERADMIN_USER and SUPERADMIN_PASSWORD_HASH:
            if check_password_hash(SUPERADMIN_PASSWORD_HASH, password):
                user = db.query(Usuario).filter_by(username=username).first()
                if not user:
                    user = Usuario(username=username, rol="superadmin", permisos="all")
                    db.add(user)
                    db.commit()
                session["usuario"] = user.username
                session["rol"] = user.rol
                return redirect("/dashboard")

        if autenticar_ldap(username, password):
            user = db.query(Usuario).filter_by(username=username).first()
            if not user:
                grupos = obtener_grupos_ldap(username)
                user = Usuario(username=username, rol=mapear_rol_por_grupo(grupos), permisos="")
                db.add(user)
            elif not user.rol or user.rol == "usuario":
                user.rol = mapear_rol_por_grupo(obtener_grupos_ldap(username))

            db.commit()
            session["usuario"] = user.username
            session["rol"] = user.rol
            return redirect("/dashboard")

        error = "Credenciales invalidas"

    return render_template("login.html", error=error)


@app.route("/dashboard")
def dashboard():
    db = get_db()
    if "usuario" not in session:
        return redirect("/")

    try:
        total_ldap = len(obtener_usuarios_ldap_cache())
        total_db = db.query(Usuario).count()
        total_admins = db.query(Usuario).filter(Usuario.rol.in_(["admin", "superadmin"])).count()
        actividad = db.query(Auditoria).order_by(Auditoria.id.desc()).limit(10).all()
        print("Hora Python:", datetime.now())
    except Exception as exc:
        print("DASHBOARD ERROR:", exc)
        total_ldap = 0
        total_db = 0
        total_admins = 0
        actividad = []

    return render_template(
        "dashboard.html",
        total_ldap=total_ldap,
        total_db=total_db,
        total_admins=total_admins,
        actividad=actividad,
        now=datetime.now(),
        notificaciones=get_unread_notifications_count(db, session.get("usuario")),
    )


@app.route("/api/dashboard")
def api_dashboard():
    db = get_db()
    try:
        actividad = db.query(Auditoria).order_by(Auditoria.id.desc()).limit(5).all()
        actividad_json = [
            {
                "admin": registro.usuario_admin,
                "accion": registro.accion,
                "usuario": registro.usuario_afectado,
                "detalle": registro.detalle,
                "fecha": serialize_datetime(registro.fecha),
            }
            for registro in actividad
        ]
        return jsonify(
            {
                "total_ldap": len(obtener_usuarios_ldap_cache()),
                "total_db": db.query(Usuario).count(),
                "total_admins": db.query(Usuario).filter(Usuario.rol.in_(["admin", "superadmin"])).count(),
                "actividad": actividad_json,
                "hora_servidor": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            }
        )
    except Exception as exc:
        print("API DASHBOARD ERROR:", exc)
        return jsonify({"error": True}), 500


@app.route("/api/reportes")
def api_reportes():
    db = get_db()

    try:

        filtros = get_report_filters(request.args)

        filtros["pagina"] = int(
            request.args.get("pagina", 1)
        )

        filtros["por_pagina"] = 8

        return jsonify(
            build_report_context(
                db,
                filtros
            )
        )

    except Exception as exc:
        print("API REPORTES ERROR:", exc)

        return jsonify({
            "error": True,
            "msg": str(exc)
        }), 500


@app.route("/usuarios")
@requiere_permiso("usuarios")
def usuarios():
    db = get_db()
    page_ldap = request.args.get("page_ldap", default=1, type=int)
    page_activos = request.args.get("page_activos", default=1, type=int)
    per_page = 7

    ldap_users_all = obtener_usuarios_ldap_cache()
    total_ldap = len(ldap_users_all)
    total_pages = max(1, ceil(total_ldap / per_page)) if total_ldap else 1
    ldap_users = ldap_users_all[(page_ldap - 1) * per_page : page_ldap * per_page]

    query = db.query(Usuario).filter(
        or_(
            func.coalesce(Usuario.permisos, "") != "",
            Usuario.rol != "usuario",
        )
    )
    total_db = query.count()
    total_pages_db = max(1, ceil(total_db / per_page)) if total_db else 1
    usuarios_db = query.offset((page_activos - 1) * per_page).limit(per_page).all()

    return render_template(
        "usuarios.html",
        ldap_users=ldap_users,
        total_pages=total_pages,
        current_page=page_ldap,
        usuarios_db=usuarios_db,
        total_db=total_db,
        total_pages_activos=total_pages_db,
        current_page_activos=page_activos,
        permisos=db.query(Permiso).all(),
        roles=db.query(Rol).all(),
        usuarios_dict={usuario.username: usuario for usuario in db.query(Usuario).all()},
        notificaciones=get_unread_notifications_count(db, session.get("usuario")),
    )


@app.route("/guardar_usuario", methods=["POST"])
def guardar_usuario():
    db = get_db()
    try:
        print("FORM:", request.form)
        username = request.form.get("username")
        rol = request.form.get("rol")
        permisos = request.form.getlist("permisos[]")

        # usuario logueado
        usuario_actual = session.get("usuario")

        admin_actual = db.query(Usuario)\
            .filter_by(username=usuario_actual)\
            .first()

        if not admin_actual:
            return jsonify({
                "status": "error",
                "msg": "Usuario inválido"
            }), 403

        if not username or not rol:
            return jsonify({"status": "error", "msg": "Datos incompletos"}), 400
        if username == "admin_root":
            return jsonify({"status": "error", "msg": "No se puede modificar el usuario principal"}), 403

        permisos_str = ",".join(permisos) if permisos else ""
        admin = session.get("usuario") or "sistema"
        user = db.query(Usuario).filter_by(username=username).first()

        # ==========================
        # SEGURIDAD DE ROLES
        # ==========================

        # Solo superadmin puede asignar superadmin
        if rol == "superadmin" and admin_actual.rol != "superadmin":
            return jsonify({
                "status": "error",
                "msg": "Solo un superadmin puede asignar ese rol"
            }), 403

        # Si el usuario destino ya es superadmin,
        # solo otro superadmin puede modificarlo
        if (
            user
            and user.rol == "superadmin"
            and admin_actual.rol != "superadmin"
        ):
            return jsonify({
                "status": "error",
                "msg": "No puedes modificar un superadmin"
            }), 403

        if user:
            user.rol = rol
            user.permisos = permisos_str
            accion = "EDITAR"
            detalle = permisos_str or rol
        else:
            user = Usuario(username=username, rol=rol, permisos=permisos_str)
            db.add(user)
            accion = "CREAR"
            detalle = permisos_str or rol

        db.commit()
        log_auditoria(admin, accion, username, detalle)

        notificacion = Notificacion(
            mensaje=f"{admin} modifico a {username}",
            usuario=admin,
            leido=0,
            fecha=datetime.now(),
        )
        db.add(notificacion)
        db.commit()

        socketio.emit(
            "nueva_notificacion",
            {
                "mensaje": f"{admin} modifico a {username}",
                "fecha": serialize_datetime(datetime.now())
            }
        )
        socketio.emit(
            "actualizar_dashboard",
            {
                "tipo": "usuarios"
            }
        )
        socketio.emit(
            "activity_feed",
            {
                "mensaje":
                    f"{admin} modificó a {username}",

                "hora":
                    datetime.now().strftime("%H:%M")
            }
        )
        return jsonify({"status": "ok"})
    except Exception as exc:
        db.rollback()
        print("ERROR API PUNTOS:", str(exc))
        return jsonify({"status": "error", "msg": str(exc)}), 500


@app.route("/api/usuario/<username>")
def api_usuario(username):
    db = get_db()
    try:
        user = db.query(Usuario).filter_by(username=username).first()
        if not user:
            return jsonify({"rol": "usuario", "permisos": []})
        permisos = [valor for valor in (user.permisos or "").split(",") if valor]
        return jsonify({"rol": user.rol or "usuario", "permisos": permisos})
    except Exception as exc:
        print("API USUARIO ERROR:", exc)
        return jsonify({"error": True}), 500


@app.route("/reportes", methods=["GET", "POST"])
def reportes():
    db = get_db()
    if "usuario" not in session:
        return redirect("/")
    filtros = get_report_filters(request.values)
    context = build_report_context(db, filtros)
    return render_template("reportes.html", **context, notificaciones=get_unread_notifications_count(db, session.get("usuario")))


@app.route("/exportar_reportes", methods=["GET", "POST"])
def exportar_reportes():
    db = get_db()
    filtros = get_report_filters(request.values)
    context = build_report_context(db, filtros)
    formato = (request.values.get("formato") or "excel").lower()
    if formato == "csv":
        return exportar_reportes_csv(context)
    return exportar_reportes_excel(context)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


@app.route("/eliminar_usuario/<username>", methods=["POST"])
def eliminar_usuario(username):

    db = get_db()

    try:

        # No tocar admin_root
        if username == "admin_root":
            return jsonify({
                "status": "error",
                "msg": "No se puede eliminar admin_root"
            }), 403

        admin = session.get("usuario")

        admin_actual = db.query(Usuario)\
            .filter_by(username=admin)\
            .first()

        user = db.query(Usuario)\
            .filter_by(username=username)\
            .first()

        if not user:
            return jsonify({
                "status": "error",
                "msg": "Usuario no encontrado"
            }), 404

        # Solo superadmin puede modificar otro superadmin
        if (
            user.rol == "superadmin"
            and admin_actual.rol != "superadmin"
        ):
            return jsonify({
                "status": "error",
                "msg": "No puedes eliminar un superadmin"
            }), 403

        # ==========================
        # QUITAR ACCESO
        # ==========================

        user.rol = "usuario"
        user.permisos = ""

        db.commit()

        socketio.emit(
            "nueva_notificacion",
            {
                "mensaje": f"{admin} quitó acceso a {username}",
                "fecha": serialize_datetime(datetime.now())
            }
        )

        socketio.emit(
            "actualizar_dashboard",
            {
                "tipo": "usuarios"
            }
        )

        socketio.emit(
            "activity_feed",
            {
                "mensaje": f"{admin} quitó acceso a {username}",
                "hora": datetime.now().strftime("%H:%M")
            }
        )

        return jsonify({"status": "ok"})

    except Exception as exc:

        db.rollback()

        print("ERROR API PUNTOS:", str(exc))

        return jsonify({
            "status": "error",
            "msg": str(exc)
        }), 500


@app.route("/api/notificaciones")
def api_notificaciones():
    db = get_db()
    try:
        usuario = session.get("usuario")
        user = db.query(Usuario).filter_by(username=usuario).first()
        if not user or user.rol == "usuario":
            return jsonify({"status": "ok", "total": 0, "data": []})

        registros = db.query(Notificacion).filter(
            Notificacion.usuario == usuario,
            Notificacion.leido == 0,
        ).order_by(Notificacion.id.desc()).limit(10).all()

        data = [
            {"id": registro.id, "mensaje": registro.mensaje, "fecha": serialize_datetime(registro.fecha)}
            for registro in registros
        ]
        return jsonify({"status": "ok", "total": len(data), "data": data})
    except Exception as exc:
        print("API NOTIFICACIONES ERROR:", exc)
        return jsonify({"status": "error", "data": []}), 500


@app.route("/api/notificaciones/leer", methods=["POST"])
def leer_notificaciones():
    db = get_db()
    usuario = session.get("usuario")
    if not usuario:
        return jsonify({"status": "error"}), 401
    try:
        db.query(Notificacion).filter(
            Notificacion.usuario == usuario,
            Notificacion.leido == 0,
        ).update({"leido": 1})
        db.commit()
        return jsonify({"status": "ok"})
    except Exception as exc:
        db.rollback()
        print("ERROR API PUNTOS:", str(exc))
        return jsonify({"status": "error", "msg": str(exc)}), 500


@app.route("/api/notificaciones/<int:notif_id>/leer", methods=["POST"])
def leer_notificacion(notif_id):
    db = get_db()
    usuario = session.get("usuario")
    try:
        notif = db.query(Notificacion).filter_by(id=notif_id, usuario=usuario).first()
        if notif:
            notif.leido = 1
            db.commit()
        return jsonify({"status": "ok"})
    except Exception as exc:
        db.rollback()
        print("ERROR API PUNTOS:", str(exc))
        return jsonify({"status": "error", "msg": str(exc)}), 500


@app.route("/api/notificaciones/clear", methods=["POST"])
def limpiar_notificaciones():
    db = get_db()
    usuario = session.get("usuario")
    try:
        db.query(Notificacion).filter(Notificacion.usuario == usuario).delete()
        db.commit()
        return jsonify({"status": "ok"})
    except Exception as exc:
        db.rollback()
        print("ERROR API PUNTOS:", str(exc))
        return jsonify({"status": "error", "msg": str(exc)}), 500


@app.route("/api/usuarios_activos")
def api_usuarios_activos():
    db = get_db()
    try:
        usuarios_con_acceso = db.query(Usuario).filter(
            or_(func.coalesce(Usuario.permisos, "") != "", Usuario.rol != "usuario")
        ).all()
        return jsonify(
            [
                {"username": usuario.username, "rol": usuario.rol or "usuario", "permisos": usuario.permisos or ""}
                for usuario in usuarios_con_acceso
            ]
        )
    except Exception as exc:
        print("API USUARIOS ACTIVOS ERROR:", exc)
        return jsonify([])


@app.route("/api/buscar_ldap")
def api_buscar_ldap():
    q = (request.args.get("q") or "").lower()
    usuarios_ldap = obtener_usuarios_ldap_cache()
    if q:
        usuarios_ldap = [usuario for usuario in usuarios_ldap if q in usuario.lower()]
    return jsonify(usuarios_ldap)


@app.route("/puntos")
def puntos():
    if "usuario" not in session:
        return redirect("/")
    return render_template("puntos.html", notificaciones=get_unread_notifications_count(get_db(), session.get("usuario")))


@app.route("/api/buscar_paciente")
def buscar_paciente():
    
    valor = request.args.get("q", "").strip()
    print("VALOR:", valor)
    if not valor.isdigit():
        return jsonify([])

    conn = None
    cur = None

    try:

        conn = oracle_pool.acquire()
        cur = conn.cursor()

        # =====================================
        # 1. Buscar por IDCLIENTE
        # =====================================

        cur.execute("""
            SELECT
                IDCLIENTE,
                NOMBRE,
                APATERNO,
                AMATERNO
            FROM SISTSIO.CLIENTES
            WHERE IDCLIENTE = :valor
        """, {
            "valor": int(valor)
        })

        row = cur.fetchone()

        # =====================================
        # 2. Si no existe, buscar por IDCITA
        # =====================================

        if not row:

            cur.execute("""
                SELECT IDCLIENTE
                FROM SISTSIO.RIOREWARDS_DETALLE
                WHERE IDCITA = :valor
            """, {
                "valor": int(valor)
            })

            prueba = cur.fetchone()

            cur.execute("""
                SELECT
                    c.IDCLIENTE,
                    c.NOMBRE,
                    c.APATERNO,
                    c.AMATERNO
                FROM SISTSIO.CLIENTES c,
                    SISTSIO.RIOREWARDS_DETALLE r
                WHERE r.IDCLIENTE = c.IDCLIENTE
                AND r.IDCITA = :valor
                AND ROWNUM = 1
            """, {
                "valor": int(valor)
            })

            row = cur.fetchone()

        if not row:
            return jsonify([])

        return jsonify([{
            "idcliente": int(row[0]),
            "nombre": f"{row[1]} {row[2]} {row[3]}",
            "saldo": get_oracle_points(int(row[0]))
        }])

    except Exception as exc:
        import traceback

        print("BUSCAR PACIENTE ERROR:")
        traceback.print_exc()

        return jsonify([])

@app.route("/api/puntos", methods=["POST"])
def modificar_puntos():
    db = get_db()
    data = request.json or {}

    try:
        idcliente = data.get("idcliente")
        idcita = data.get("idcita")
        if idcita not in [None, "", "null"]:
            idcita = str(idcita)
        else:
            idcita = None

        tipo = data.get("tipo")

        try:
            cantidad = int(data.get("cantidad"))
        except Exception:
            return jsonify({"status": "error", "msg": "Cantidad invalida"}), 400

        if "usuario" not in session:
            return jsonify({"status": "error", "msg": "Sesion expirada"}), 401
        if not idcliente or not tipo:
            return jsonify({"status": "error", "msg": "Datos incompletos"}), 400
        if cantidad <= 0:
            return jsonify({"status": "error", "msg": "Cantidad invalida"}), 400
        if tipo not in {"suma", "resta"}:
            return jsonify({"status": "error", "msg": "Tipo invalido"}), 400
        categoria = data.get("categoria", "ajuste")

        # 🔒 solo admin/superadmin pueden quitar reseñas
        usuario_actual = session.get("usuario")

        usuario_db = db.query(Usuario).filter_by(
            username=usuario_actual
        ).first()

        rol_usuario = usuario_db.rol if usuario_db else "usuario"

        if (
            tipo == "resta"
            and categoria == "reseña"
            and rol_usuario not in ["admin", "superadmin"]
        ):
            return jsonify({
                "status": "error",
                "msg": "No tienes permisos para retirar puntos de reseña"
            }), 403


        # 💰 Validación saldo
        saldo_actual = get_operational_points(db, idcliente)

        if tipo == "resta" and cantidad > saldo_actual["total"]:
            return jsonify({
                "status": "error",
                "msg": "Saldo insuficiente",
                "saldo_actual": saldo_actual["total"]
            }), 400

        if tipo == "suma" and categoria == "reseña":

            if paciente_tiene_resena_oracle(idcliente):

                return jsonify({
                    "status": "error",
                    "msg": "Este paciente ya tiene una reseña aplicada"
                }), 400
        # 💾 Guardar
        nuevo = MovimientoPuntos(
            paciente_id=str(idcliente),
            idcita=idcita,
            tipo=tipo,
            cantidad=cantidad,
            motivo=categoria,
            origen="manual",
            usuario_admin=session.get("usuario")
        )

        try:
            db.add(nuevo)

            if tipo == "suma" and categoria == "reseña":

                insertar_puntos_oracle(
                    idcliente=idcliente,
                    idcita=idcita,
                    cantidad=cantidad,
                    categoria=categoria,
                    usuario=session.get("usuario")
                )
            if tipo == "resta" and categoria == "reseña":

                insertar_cancelacion_resena_oracle(
                    idcliente=idcliente,
                    idcita=idcita,
                    cantidad=cantidad,
                    usuario=session.get("usuario")
                )    
            db.commit()
            if tipo == "suma" and categoria == "reseña":
                print(
                    f"RESEÑA APLICADA -> cliente={idcliente} cita={idcita}"
                )

            if tipo == "resta" and categoria == "reseña":
                print(
                    f"RESEÑA CANCELADA -> cliente={idcliente} cita={idcita}"
                )
            db.refresh(nuevo)
            # 🔔 crear notificación
            mensaje = (
                f"{session.get('usuario')} agregó {cantidad} puntos"
                if tipo == "suma"
                else f"{session.get('usuario')} retiró {cantidad} puntos"
            )

            notif = Notificacion(
                mensaje=mensaje,
                usuario=session.get("usuario"),
                leido=0,
                fecha=datetime.now()
            )

            db.add(notif)
            db.commit()

            # 🔥 socket tiempo real
            socketio.emit(
                "nueva_notificacion",
                {
                    "mensaje": mensaje,
                    "fecha": serialize_datetime(datetime.now())
                }
            )

            socketio.emit(
                "actualizar_dashboard",
                {
                    "tipo": "puntos"
                },
                
            )
            socketio.emit(
                "nuevo_movimiento",
                {
                    "fecha":
                        datetime.now().strftime("%Y-%m-%d"),

                    "puntos":
                        cantidad,

                    "admin":
                        session.get("usuario"),

                    "tipo":
                        tipo
                }
            )

            socketio.emit(
                "activity_feed",
                {
                    "mensaje":
                        (
                            f"{session.get('usuario')} agregó {cantidad} puntos"
                            if tipo == "suma"
                            else f"{session.get('usuario')} retiró {cantidad} puntos"
                        ),
                    "hora":
                        datetime.now().strftime("%H:%M")
                }
            )
            nuevo_total = sum(
                (
                    mov.cantidad
                    if mov.tipo == "suma"
                    else -mov.cantidad
                )
                for mov in db.query(
                    MovimientoPuntos
                ).all()
            )

            total_movimientos = db.query(
                MovimientoPuntos
            ).count()
            
            socketio.emit(
                "kpi_update",
                {
                    "puntos": nuevo_total,
                    "movimientos": total_movimientos
                }
            )

        except IntegrityError:
            db.rollback()
            return jsonify({
                "status": "error",
                "msg": "Esta cita ya tiene reseña aplicada (bloqueo BD)"
            }), 400

        # 🔄 limpiar cache
        CACHE_PUNTOS.pop(str(idcliente), None)

        saldo_nuevo = get_operational_points(db, idcliente)

        saldo_nuevo = get_operational_points(db, idcliente)

        return jsonify({
            "status": "ok",
            "saldo": saldo_nuevo,
            "movimiento": serialize_movimiento(nuevo)
        })
    except Exception as exc:
        db.rollback()
        print("ERROR API PUNTOS:", str(exc))
        return jsonify({"status": "error", "msg": str(exc)}), 500


@app.route("/api/historial/<idcliente>")
def historial(idcliente):
    return jsonify(
        obtener_historial_oracle(idcliente)
    )

@app.route("/api/paciente/<idcliente>")
def api_paciente(idcliente):

    db = get_db()

    try:

        historial = obtener_historial_oracle(idcliente)
        saldo = get_operational_points(db, idcliente)

        return jsonify({
            "puntos": saldo["total"],
            "saldo": saldo,
            "historial": historial,
            "tiene_resena": tiene_resena_oracle(idcliente)
        })

    except Exception as exc:

        print("ERROR API PACIENTE:", exc)

        return jsonify({
            "error": True,
            "msg": str(exc)
        }), 500

def get_puntos_cache(idcliente):
    return get_oracle_points(idcliente)


@app.route("/api/tiene_resena/<idcliente>")
def api_tiene_resena(idcliente):
    return jsonify({"tiene": tiene_resena(idcliente)})


@app.route("/api/grafica/puntos")
def api_grafica_puntos():
    db = get_db()
    filtros = get_report_filters(request.args)
    context = build_report_context(db, filtros)
    chart = context["chart_puntos_fecha"]
    return jsonify(
        [
            {"dia": chart["labels"][index], "total": chart["valores"][index]}
            for index in range(len(chart["labels"]))
        ]
    )


@app.route("/api/dashboard/kpi")
def api_dashboard_kpi():
    db = get_db()
    filtros = get_report_filters(request.args)
    context = build_report_context(db, filtros)
    return jsonify(
        {
            "total": context["kpis"]["puntos_netos"],
            "hoy": context["kpis"]["movimientos_hoy"],
            "movimientos": context["kpis"]["movimientos_puntos"],
            "auditoria": context["kpis"]["auditoria"],
        }
    )


@app.route("/api/reportes/detalle")
def detalle_reporte():
    db = get_db()
    tipo = (request.args.get("tipo") or "").lower()

    if tipo in {"admin", "usuario", "superadmin"}:
        usuarios = db.query(Usuario).filter(Usuario.rol == tipo).all()
        return jsonify([usuario.username for usuario in usuarios])

    if tipo in {"suma", "resta"}:
        movimientos = db.query(MovimientoPuntos).filter(
            MovimientoPuntos.tipo == tipo
        ).order_by(MovimientoPuntos.fecha.desc()).limit(20).all()
        return jsonify([serialize_movimiento(movimiento) for movimiento in movimientos])

    return jsonify([])

@app.route("/transcripcion")
def transcripcion():
    return render_template("transcripcion.html")

@app.route("/api/transcripcion/estado")
def estado_transcripcion():

    conn = None
    cursor = None

    try:

        conn = oracle_pool.acquire()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                CLAVE,
                MODO,
                IDCITA,
                FECHA_MOD,
                USUARIO_MOD
            FROM SISTSIO.V_SEMAFORO_UNIFICADO_ESTADO
        """)

        row = cursor.fetchone()

        if not row:

            return jsonify({
                "status": "success",
                "modo": None
            })

        return jsonify({
            "status": "success",
            "clave": row[0],
            "modo": row[1],
            "idcita": row[2],
            "fecha_mod": str(row[3]) if row[3] else None,
            "usuario_mod": row[4]
        })

    except Exception as e:

        print("ERROR ESTADO:", e)

        return jsonify({
            "status": "error",
            "msg": str(e)
        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

@socketio.on("connect")
def socket_connect():

    try:

        username = session.get("usuario")

        if username:

            room = f"user_{username}"

            join_room(room)

            usuarios_online[username] = True

            emit(
                "usuarios_online",
                list(usuarios_online.keys()),
                broadcast=True
            )

    except Exception as e:

        print("SOCKET CONNECT ERROR:", e)

@socketio.on("disconnect")
def socket_disconnect():

    try:

        username = session.get("usuario")

        if username:

            room = f"user_{username}"

            leave_room(room)

            usuarios_online.pop(
                username,
                None
            )

            emit(
                "usuarios_online",
                list(usuarios_online.keys()),
                broadcast=True
            )

    except Exception as e:

        print(
            "SOCKET DISCONNECT ERROR:",
            e
        )


if __name__ == "__main__":
    socketio.run(
    app,
    host="0.0.0.0",
    port=5000,
    debug=False,
    use_reloader=False
)
